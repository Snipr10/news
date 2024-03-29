import asyncio
import time

import requests
from newspaper import Article
import re
import telebot
from telethon import TelegramClient
from telebot import types
from telebot.types import InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import load_workbook
from bs4 import BeautifulSoup, NavigableString

bot = telebot.TeleBot('2081961920:AAERnhcfR-cEjw84VKwLG1RGc3Ra22G0v3k')

MESSAGE_NEW = "Отправь ссылку на статью в ответ на это сообщение"
MESSAGE_KEY = "Отправь ссылку на статью и текст ошибки в ответ на это сообщение"
MESSAGE_ERROR_URL = "Не могу найти ссылку, в сообщении должна быть одна ссылка"
MESSAGE_ERROR = "Что-то пошло не так"
MESSAGE_NOT_FOUND_TASK = "Не мог найти таску, похожие:"
MESSAGE_RESET_TASKS = "Таски сброшены"
MESSAGE_ERROR_SAVE = "Ошибка сохранена"
ERROR_CHAT = "1974656292"
URL_PATTERN = r'/^(https?:\/\/)?([\da-z\.-]+)\.([a-z\.]{2,6})([\/\w \.-]*)*\/?$/'

URL_TG_API = "http://194.50.24.4:8000/api/"


def gen_markup():
    markup = InlineKeyboardMarkup()
    markup.row_width = 3
    markup.add(
        InlineKeyboardButton("Распарсить статью", callback_data="news"),
        InlineKeyboardButton("Сообщить об ошибке", callback_data="error"),

    )
    return markup


def gen_markup_message(text):
    markup = InlineKeyboardMarkup()
    markup.row_width = 3
    markup.add(
        InlineKeyboardButton(u'\u2705', callback_data="ok"),
        InlineKeyboardButton(u'\u274c', callback_data="bad"),

    )
    return markup


@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    try:
        chat_id = call.message.chat.id
        if call.data == "news":
            bot.send_message(chat_id,
                             text=MESSAGE_NEW, reply_markup=types.ForceReply())

        elif call.data == "error":
            bot.send_message(chat_id,
                             text=MESSAGE_KEY, reply_markup=types.ForceReply())

        elif call.data == "ok":
            try:
                urls = get_url(call.message.json['entities'][-1]['url'])
            except Exception:
                urls = get_url(
                    re.search("(?P<url>https?://[^\s]+)", call.message.json["text"]).group("url").replace(")", ""))
            res = requests.post(URL_TG_API + "add_parsing_site", json={"urls": urls})
            if res.ok:
                bot.send_message(chat_id, text=f"Сайт добавлен для парсинга: {urls[0]}")
            else:
                bot.send_message(chat_id, text="не могу добавить урл, поытайтесь позже")

        elif call.data == "bad":
            # bot.send_message(ERROR_CHAT, call.message.json['entities'][0]['url'])
            try:
                url = call.message.json['entities'][-1]['url']
            except Exception:
                url = re.search("(?P<url>https?://[^\s]+)", call.message.json["text"]).group("url").replace(")", "")

            bot.send_message(chat_id,
                             text=f"Опишите проблему [Link]({url}) \n",
                             parse_mode="Markdown",
                             reply_markup=types.ForceReply())

        @bot.message_handler()
        def proc_reply(message):
            try:
                if message.reply_to_message.text == MESSAGE_NEW:
                    send_message_new(message)
                elif message.reply_to_message.text == MESSAGE_KEY:
                    send_message_error(message)
                elif "Опишите проблему" in message.reply_to_message.text:
                    bot.forward_message(ERROR_CHAT, message.chat.id, message.message_id)
                    bot.send_message(ERROR_CHAT, message.reply_to_message.json['entities'][-1]['url'])
                    try:
                        try:
                            urls = get_url(message.reply_to_message.json['entities'][-1]['url'])
                        except Exception:
                            urls = get_url(
                                re.search("(?P<url>https?://[^\s]+)", call.message.json["text"]).group("url").replace(
                                    ")", ""))

                        res = requests.post(URL_TG_API + "check_parsing_site", json={"urls": urls})
                        if int(res.text) > 0:
                            res = requests.post(URL_TG_API + "deactivate_parsing_site", json={"urls": urls})
                            bot.send_message(message.chat.id,
                                             "Этот сайт раньше работал правильно, сейчас мы деактивировали его")

                        else:
                            bot.send_message(message.chat.id, MESSAGE_ERROR_SAVE)

                    except Exception:
                        bot.send_message(message.chat.id, MESSAGE_ERROR_SAVE)

            except Exception:
                bot.send_message(message.chat.id, "Отправьте сообщение ответом")
    except Exception:
        pass


def get_url(ful_url):
    from urllib.parse import urlparse
    parse_url = urlparse(ful_url)
    url = f"{parse_url.scheme}://{parse_url.hostname}"
    return [url, url + "/"]


@bot.message_handler(commands=['start'])
def send_welcome(message):
    bot.reply_to(message,
                 text=f'Привет, я буду помогать тебе со статьями, перейди в меню: /menu')


@bot.message_handler(commands=['menu'])
def message_handler(message):
    bot.send_message(message.chat.id, "Выбери действие", reply_markup=gen_markup())


@bot.message_handler(commands=['statistic'])
def message_handler(message):
    wb = load_workbook(filename='news_text_bot.xlsx')
    sheet = wb['Лист1']
    ws = wb.active
    i = 2
    n = True
    for m in client.iter_messages("news_text_bot", reverse=True):
        if n:
            ws["B%s" % i] = m.text
            n = False
        else:
            ws["A%s" % i] = m.text
            n = True
            i += 1
    name = "errors-" + str(time.time()).replace(".", "") + ".xlsx"
    wb.save(name)

    f = open(name, "rb")
    bot.send_document(message.chat.id, f)


def send_message_error(message):
    try:
        if not check_url(message.text):
            bot.send_message(message.chat.id, MESSAGE_ERROR_URL)
        else:
            bot.send_message(ERROR_CHAT, message.text)

            bot.send_message(message.chat.id, MESSAGE_ERROR_SAVE)

    except Exception:
        bot.send_message(message.chat.id, "Не могу записать ошибку")


def chunkstring(string, length):
    return (string[0 + i:length + i] for i in range(0, len(string), length))


def send_message_new(message):
    try:
        if not check_url(message.text):
            bot.send_message(message.chat.id, MESSAGE_ERROR_URL)
        else:
            article_title, article_meta_description, article_publish_date, text = get_article_data(message.text)
            text = re.sub("\n\n+", "\n\n", text)
            message_text = ""
            message_text += "Title: \n" + (article_title or "") + "\n"
            message_text += "Meta: \n" + (article_meta_description or "") + "\n"
            message_text += "Text: \n" + text + "\n"
            message_text += "Date: \n" + str(article_publish_date or "") + "\n"
            message_text += f"[Link]({message.text}) \n"

            messages_list = chunkstring(message_text, 4095)
            m_list = list(messages_list)
            for i in range(len(m_list)):
                if i == len(m_list) - 1:
                    if message.text not in m_list[i]:
                        m_list[i] = message.text
                    bot.send_message(message.chat.id, m_list[i], reply_markup=gen_markup_message(message.text),
                                     parse_mode="Markdown")
                else:

                    bot.send_message(message.chat.id, m_list[i],
                                     parse_mode="Markdown")

    except Exception as e:
        print(e)
        bot.send_message(message.chat.id, MESSAGE_ERROR)


URL_DICT = {
    "https://infoneva.ru/": {"title": ["title", {}], "text": ["div", {"class": "text-content"}]},
    "https://www.ntv.ru/": {"title": ["h1", {"itemprop": "headline"}], "text": ["div", {"class": "inpagebody"}]},
    "https://peterburg2.ru/": {"title": ["h1", {}], "text": ["span", {"class": "article-content"}],
                               "meta": ["p", {"class": "article-content"}]},
    # "https://lenta.ru/": {"title": ["h1", {}], "text": ["div", {"class": "topic-body"}],
    #                            "meta": ["div", {"class*=topic-header__title-yandex"}]},
    # "https://bloknot.ru/": {"title": ["h1", {}], "text": ["div", {"class": "article__content"}]}
    "https://spb.dixinews.ru/": {"title": ["h1", {}], "text": ["div", {"class": "entry-content"}], "wholetext": True},
    "https://ria.ru/": {"title": ["div", {"class": "article__title"}], "text": ["div", {"class": "article__body"}],
                        "meta": ["h1", {"class": "article__second-title"}]},
    "https://vedomosti-spb.ru/": {"title": ["h1", {"class": "article-headline__title"}],
                                  "text": ["div", {"class": "article-boxes-list article__boxes"}],
                                  "meta": ["div", {"class": "article-authors__info"}]},
    "https://live24.ru/": {"title": ["h1", {}],
                           "text": ["div", {
                               "class":  re.compile('.*panel uk-text-large maintext.*')}],
                           "is_last": True
                           },
    # "https://www.sobaka.ru/": {"title": ["h1", {"itemprop": "headline name"}],
    #                            "text": ["div", {"itemprop": "articleBody"}],
    #                            },
    "https://www.flashnord.com/": {"title": ["h1", {"class": "entry-title"}],
                                   "text": ["div", {"class": "entry-content"}],
                                   },
    "https://www.interfax-russia.ru/": {"title": ["div", {"itemprop": "headline"}],
                                        "text": ["div", {"itemprop": "articleBody"}],
                                        },
    "https://www.rtr.spb.ru/": {"title": ["font", {"class": "base"}],
                                        "text": ["p", {"align": "justify"}],
                                "decoder": "windows-1251", "manual":True
                                        },
    "https://spb.aif.ru/": {"title": ["h1", {}], "text": ["div", {"class": "article_text"}], "p": True, "next": True
                             },
    "https://galernayas.ru/": {"title": ["h2", {}], "text": ["div", {"style": "text-align:justify"}], "p": True,
                            },
    "https://78.ru/": {"title": ["h1", {}], "text": ["div", {"class": "publication__body"}],
                               "meta": ["div", {"class": "article__summary"}]},
    "https://novayagazeta.spb.ru/": {"title": ["h1", {}], "text": ["div", {"class": "article"}],
                               },
    "https://www.interessant.ru/": {"title": ["h1", {}], "text": ["div", {"class": "text"}],
                       "meta": ["h2", {}], "is_last": True},
    "https://www.lenpravda.ru/": {"title": ["div", {"class":"razdeltitle1"}], "text": ["div", {"class": "bodytext"}],
                                  "is_last": True},
    "https://info24.ru/": {"title": ["h1",{}], "meta": ["h3", {}], "text": ["div", {"class": "material-body"}],
                                  "is_last": True},
    "http://www.assembly.spb.ru/": {"title": ["h2", {}], "text": ["article", {}],
                           "is_last": True},
    "https://govoritmoskva.ru/": {"title": ["h1", {}], "text": ["div", {"class": "textContent"}],
                                    "is_last": True},
    "https://smotrim.ru/": {"title": ["header", {"class":"article-main-item__header"}], "text": ["div", {"class": "article-main-item__body"}],
                                  "meta":["div", {"class":"article-main-item__anons"}],
                                  "is_last": True},
    "https://hudoznikov.ru/": {"title": ["h2", {}], "text": ["div", {"style": "text-align:justify"}],
                                  "is_last": True},
    "https://dorinfo.ru/": {"title": ["h1", {}],
                            "text": ["div", {"class": "fulltext"}],
                            "is_last": True},
    "https://www.flashnord.com/": {"title": ["h1", {}],
                            "text": ["div", {"class": "entry-content"}],
                            "is_last": True},
    "https://www.vedomosti.ru/": {"title": ["h1", {}],
                            "text": ["div", {"class": "article-boxes-list article__boxes"}],
                            "meta": ["em", {}],},
    # "https://novayagazeta.ru/": {"title": ["h1", {}],
    #                               "text": ["div", {"id": "materialBlock_0"}],
    #                             },
    "https://mir24.tv/": {"title": ["h1", {}],
                                 "text": ["div", {"class": "article-content"}], "p": True, "next": True
                                 },
    "https://mos.news/": {"title": ["h1", {}],
                          "text": ["div", {"class": "detail_text_container"}],  "decoder": "windows-1251",  "next": True
                          },
    "https://delta.news/": {"title": ["h5", {"class":"white-text grey darken-2"}],
                          "text": ["article", {"class": "card"}], "next": True,
                            "delete_title": True
                          },
    "https://spb.octagon.media/": {"title": ["h1", {}],
                            "text": ["article", {}], "p": True,
                            "delete_title": True
                            },
    "https://bloknot.ru/": {"title": ["h1", {}],
                                   "text": ["div", {"class":"article__content"}], "p": True,
                                   "delete_title": True, "next": True
                                   },
    "https://www.sobaka.ru/": {"title": ["h1", {}],
                                   "text": ["div", {"class":"b-post-blocks"}], "p": True,
                                   }
}


def _get_page_data(url):
    for k in URL_DICT.keys():
        if k in url:
            try:
                post = requests.get(url)
            except Exception:
                post = requests.get(url, headers={
                    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.75 Safari/537.36'
                })
            if URL_DICT.get(k).get("decoder"):
                soup = BeautifulSoup(post.content.decode(URL_DICT.get(k).get("decoder")))
            else:
                soup = BeautifulSoup(post.text, 'html.parser')
            article_title = soup.find(name=URL_DICT.get(k).get("title")[0], attrs=URL_DICT.get(k).get("title")[1]).text
            text = ""
            try:
                if "meta" in URL_DICT.get(k).keys():
                    for c in soup.find(name=URL_DICT.get(k).get("meta")[0],
                                       attrs=URL_DICT.get(k).get("meta")[1]).contents:
                        try:
                            if c.text and c.text.strip():
                                text += c.text + "\r\n <br> "
                        except Exception:
                            pass
            except Exception:
                pass
            if URL_DICT.get(k).get("wholetext"):
                text += soup.find(name=URL_DICT.get(k).get("text")[0], attrs=URL_DICT.get(k).get("text")[1]).text
            else:
                soup_all = soup.find_all(name=URL_DICT.get(k).get("text")[0], attrs=URL_DICT.get(k).get("text")[1])
                if URL_DICT.get(k).get("is_last", False):
                    soup_cont = soup_all[-1]
                else:
                    soup_cont = soup_all[0]
                text = ""

                if URL_DICT.get(k).get("manual", False):
                    for c in soup_cont.contents[0].contents:
                        if isinstance(c, NavigableString) or len(c.attrs) == 0:
                            text += str(c)
                        else:
                            break
                else:
                    if URL_DICT.get(k).get("p", False):
                        for c in soup_cont.find_all("p"):
                            try:
                                if URL_DICT.get(k).get("next", False):
                                    if c.next and c.next.strip():
                                        text += re.sub("\n+", "\n", c.next.strip()) + "\r\n <br> "
                                else:
                                    if c.text and c.text.strip():
                                        text += re.sub("\n+", "\n", c.text.strip()) + "\r\n <br> "
                            except Exception:
                                pass
                    else:
                        for c in soup_cont.contents:
                            try:
                                if c.text and c.text.strip():
                                    text += re.sub("\n+", "\n", c.text.strip()) + "\r\n <br> "
                            except Exception:
                                pass
            if URL_DICT.get(k).get("delete_title", False):
                try:
                    text = text.replace(article_title.strip(), "").strip()
                except Exception:
                    pass
            return article_title, text
    return "", ""


def get_article_data(url):
    article_title = ""
    article_meta_description = ""
    article_publish_date = None
    text = ""
    article_title, text = _get_page_data(url)

    if article_title == "" and text == "":
        article = Article(url)
        article.download()
        article.parse()
        title = article.title
        text = article.text or ""
        text_first_part = text[:len(title) * 3]
        if title in text_first_part:
            if len(text_first_part.replace(title + "\n", "")) != len(text_first_part):
                text = text.replace(title + "\n", "", 1)
            else:
                try:
                    first_par = text.split('\n')[0]
                    first_par_without_par = first_par.replace(title, "")
                    if "." not in first_par_without_par and "!" not in first_par_without_par and "?" not in first_par_without_par:
                        text = text.replace(title, "", 1)
                except Exception:
                    pass
        article_title = article.title
        article_meta_description = article.meta_description
        article_publish_date = article.publish_date
    return article_title, article_meta_description, article_publish_date, text


def check_url(text):
    urls = re.findall(r'http(?:s)?://\S+', text)
    return len(urls) == 1


def start_bot():
    try:
        while True:
            bot.polling(none_stop=True, timeout=100, long_polling_timeout=10000000)
    except Exception:
        start_bot()


if __name__ == '__main__':
    # article_title, text = _get_page_data(
    #     "https://topspb.tv/news/2021/10/22/dva-sezda-s-primorskogo-puteprovoda-uluchshat-dorozhnuyu-situaciyu-na-severo-zapade-peterburga/")

    phone = "79910422683"
    client = TelegramClient(phone, 8296076, "eb69d92ebb65d72cbb8366ffb3ce7f0d")
    client.start(phone)
    start_bot()
